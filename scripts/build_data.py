#!/usr/bin/env python3
"""
HOSQ dashboards — data pipeline.
xlsx (v7 enriched) -> versioned static JSON in ../data/.

Honesty rules (locked with the principal):
  - "GAP:notion_missing" and empty -> null (no data), never 0.
  - jobs_created/jobs_paid/partners_international: a 0 means UNFILLED, not measured -> null + flag.
  - audience_reach_online 2,000,000 belongs to the Notations Lab flagship only -> quarantined.
  - events/team/partners DO NOT join the 47 projects (all project_id="nl"); they describe one
    programme -> emitted as a scoped `notations_lab_case`, never as portfolio aggregates.
  - A-J framework has NO numeric scores -> we publish EVIDENCE COVERAGE (fill rate), labelled as such.
  - Public build strips PII (emails, telegram handles, contact persons).

Usage:  python scripts/build_data.py [path/to/hosq_project_profile_v7_enriched.xlsx]
"""
import json, sys, datetime, re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = "/Users/kola/Library/Containers/Mail.Ru.DiskO.as/Data/Disk-O.as.mounts/kolaootro@mail.ru-525796781775651729/nikolai/HOSQ/Claude-data-metrics-normalisatoion/hosq_project_profile_v7_enriched.xlsx"
SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_XLSX)
OUT = ROOT / "data"
(OUT / "projects").mkdir(parents=True, exist_ok=True)

# Fields where a literal 0 means "unfilled default", not a measured zero.
UNVERIFIED_ZERO = {"jobs_created_count", "jobs_paid_count", "partners_international_count"}
# Quarantine: single self-reported online figure tied to the nl flagship.
ONLINE_QUARANTINE_THRESHOLD = 1_000_000
PII_TEAM = {"email", "telegram"}
PII_PARTNERS = {"contact_person", "contact_email"}

INT_FIELDS = {"project_edition_year","duration_days","team_size_total","participants_count_onsite",
    "partners_international_count","partners_local_count","jobs_created_count","jobs_paid_count",
    "mentors_count","audience_reach_onsite","audience_reach_online"}
FLOAT_FIELDS = {"budget_total_usd"}
ARRAY_FIELDS_PIPE = {"project_format_tags","team_leads_list","team_disciplines_list","languages_offered",
    "multilingual_outputs_list","mous_or_agreements_list","community_partners_list","coproductions_list",
    "methods_implemented_list","skills_transferred_list","research_artefacts_list","curatorial_texts_list",
    "tech_prototypes_list","awards_list","distribution_channels","media_coverage_list"}

def is_gap(v):
    return v is None or (isinstance(v, str) and (v.strip() == "" or v.strip().startswith("GAP")))

def clean(v):
    return None if is_gap(v) else v

def scrub_pii(s):
    """Remove bare @handles from free-text (principal asked for 'no contacts')."""
    if not isinstance(s, str): return s
    return re.sub(r"(?<!\w)@[A-Za-z0-9_]{3,}", "", s).replace("  ", " ").strip()

def split_arr(v):
    if is_gap(v): return []
    s = str(v)
    parts = re.split(r"[|\n]", s)
    return [scrub_pii(p.strip()) for p in parts if p.strip()]

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---- schema: column_id -> {data_type, metric_codes[]} ----
schema_rows = list(wb["schema"].iter_rows(values_only=True))
sch_hdr = [str(c) for c in schema_rows[0]]
ci = sch_hdr.index("column_id"); cc = sch_hdr.index("hosq_metric_codes"); ct = sch_hdr.index("data_type")
field_meta = {}
field_to_subindices = {}
for r in schema_rows[1:]:
    cid = r[ci]
    if not cid: continue
    codes = [c.strip() for c in str(r[cc]).split(",") if c and c.strip()] if r[cc] else []
    field_meta[cid] = {"type": r[ct]}
    field_to_subindices[cid] = codes

# ---- metrics legend: vectors + subindices ----
legend_rows = list(wb["metrics_legend"].iter_rows(values_only=True))
lh = [str(c) for c in legend_rows[0]]
VEC_ORDER = []
vectors = {}
for r in legend_rows[1:]:
    vc, vn, sc, ss, df = r[0], r[1], r[2], r[3], r[4]
    if not vc: continue
    if vc not in vectors:
        vectors[vc] = {"vector": vc, "name": vn, "subindices": []}
        VEC_ORDER.append(vc)
    vectors[vc]["subindices"].append({"code": sc, "short": ss, "definition": df})

# evidence fields per vector (fields whose metric codes start with that vector letter)
vector_fields = {v: [] for v in VEC_ORDER}
for field, codes in field_to_subindices.items():
    hit = set(c[0] for c in codes if c)
    for v in hit:
        if v in vector_fields:
            vector_fields[v].append(field)

# ---- projects (wide) ----
prj_rows = list(wb["projects"].iter_rows(values_only=True))
hidx = next(i for i, r in enumerate(prj_rows) if r[1] == "project_id")
header = list(prj_rows[hidx])            # column_ids, index 1.. (index 0 = meta col)
data_rows = [r for r in prj_rows[hidx + 1:] if r[1]]

def field_filled(field, raw):
    """For coverage: is there real evidence in this field?"""
    if is_gap(raw): return False
    if field in UNVERIFIED_ZERO and raw == 0: return False
    if isinstance(raw, str) and not raw.strip(): return False
    return True

def coerce(field, raw):
    v = clean(raw)
    if v is None: return None
    if field in UNVERIFIED_ZERO and v == 0: return None
    if field in INT_FIELDS:
        try: return int(float(v))
        except: return None
    if field in FLOAT_FIELDS:
        try: return float(v)
        except: return None
    if field in ARRAY_FIELDS_PIPE:
        return split_arr(raw)
    return v

projects = []
for r in data_rows:
    rec = {}
    for col in range(1, len(header)):
        field = header[col]
        if not field: continue
        rec[field] = r[col]
    pid = rec.get("project_id")
    if not pid: continue

    # vector coverage (evidence fill-rate per vector) — labelled as documentation completeness, NOT score
    vcov = {}
    for v in VEC_ORDER:
        flds = vector_fields[v]
        if not flds: vcov[v] = None; continue
        filled = sum(1 for f in flds if field_filled(f, rec.get(f)))
        vcov[v] = round(filled / len(flds), 3)

    online_raw = coerce("audience_reach_online", rec.get("audience_reach_online"))
    online_quarantined = online_raw is not None and online_raw >= ONLINE_QUARANTINE_THRESHOLD

    def isodate(v):
        v = clean(v)
        if v is None: return None
        s = str(v)[:10]
        return s if len(s) == 10 and s[4] == '-' else None
    sd = isodate(rec.get("start_date")); ed = isodate(rec.get("end_date"))
    media_list = coerce("media_coverage_list", rec.get("media_coverage_list")) or []
    bud = coerce("budget_total_usd", rec.get("budget_total_usd"))
    aud = coerce("audience_reach_onsite", rec.get("audience_reach_onsite"))

    summary = {
        "project_id": pid,
        "slug": clean(rec.get("project_slug")) or pid,
        "title": clean(rec.get("project_title")),
        "type": clean(rec.get("project_type")),
        "status": clean(rec.get("project_status")),
        "year": coerce("project_edition_year", rec.get("project_edition_year")),
        "stream": clean(rec.get("hosq_stream")),
        "city": clean(rec.get("primary_city")),
        "country": clean(rec.get("primary_country")),
        "venue": clean(rec.get("venue_name_primary")),
        "venue_type": clean(rec.get("venue_type")),
        "url": clean(rec.get("project_url")),
        "start_date": sd, "end_date": ed, "start_month": sd[:7] if sd else None,
        "duration_days": coerce("duration_days", rec.get("duration_days")),
        "tags": coerce("project_format_tags", rec.get("project_format_tags")),
        "media_count": len(media_list),
        "cost_per_attendee": round(bud / aud, 1) if (bud is not None and aud not in (None, 0)) else None,
        "languages": coerce("languages_offered", rec.get("languages_offered")),
        "short_description": clean(rec.get("short_description")),
        "budget_usd": coerce("budget_total_usd", rec.get("budget_total_usd")),
        "audience_onsite": coerce("audience_reach_onsite", rec.get("audience_reach_onsite")),
        "audience_online": None if online_quarantined else online_raw,
        "participants_onsite": coerce("participants_count_onsite", rec.get("participants_count_onsite")),
        "partners_local": coerce("partners_local_count", rec.get("partners_local_count")),
        "partners_international": coerce("partners_international_count", rec.get("partners_international_count")),
        "team_size": coerce("team_size_total", rec.get("team_size_total")),
        "vector_coverage": vcov,
        "flags": {
            "budget_missing": coerce("budget_total_usd", rec.get("budget_total_usd")) is None,
            "intl_partners_unverified": rec.get("partners_international_count") in (0, None) or is_gap(rec.get("partners_international_count")),
            "online_quarantined": online_quarantined,
        },
    }

    # full per-project detail (narratives) for project.html
    detail = dict(summary)
    NARR = ["full_description","curatorial_concept","target_audience_description","artistic_research_questions",
            "team_competence_narrative","internationality_narrative","engagement_narrative","outcomes_list",
            "collaborations_formed","ecosystem_continuity_narrative","post_effect_continuation_narrative",
            "heritage_engagement_narrative","site_specific_rationale_narrative","accessibility_physical",
            "accessibility_economic","accessibility_digital","budget_narrative","financial_innovation_narrative",
            "market_integration_narrative","risk_assessment_narrative","environmental_narrative",
            "skills_transferred_list","media_coverage_list","awards_list","coproductions_list","methods_implemented_list"]
    detail["narratives"] = {}
    for f in NARR:
        if is_gap(rec.get(f)): continue
        val = coerce(f, rec.get(f))
        detail["narratives"][f] = [scrub_pii(x) for x in val] if isinstance(val, list) else scrub_pii(val)
    detail["subindex_evidence"] = {}
    for v in VEC_ORDER:
        for si in vectors[v]["subindices"]:
            code = si["code"]
            evfields = [f for f, cs in field_to_subindices.items() if code in cs]
            has = any(field_filled(f, rec.get(f)) for f in evfields)
            detail["subindex_evidence"][code] = has

    projects.append(summary)
    (OUT / "projects" / f"{summary['slug']}.json").write_text(
        json.dumps(detail, ensure_ascii=False, indent=2))

# ---- portfolio aggregates (always with denominators) ----
def covered_sum(field_key):
    vals = [p[field_key] for p in projects if p[field_key] is not None]
    return {"value": round(sum(vals), 2) if vals else None, "n_covered": len(vals), "n_total": len(projects)}

from collections import Counter
def group_count(key):
    c = Counter((p[key] or "—") for p in projects)
    return dict(sorted(c.items(), key=lambda x: -x[1]))

distinct_langs = sorted({l for p in projects for l in (p["languages"] or [])})
distinct_disc = None  # disciplines live only in nl team table — not portfolio-wide

portfolio_vcov = {}
for v in VEC_ORDER:
    vals = [p["vector_coverage"][v] for p in projects if p["vector_coverage"][v] is not None]
    portfolio_vcov[v] = {"name": vectors[v]["name"],
                         "coverage": round(sum(vals)/len(vals), 3) if vals else None,
                         "n_fields": len(vector_fields[v])}

aggregates = {
    "project_count": len(projects),
    "by_type": group_count("type"),
    "by_stream": group_count("stream"),
    "by_year": group_count("year"),
    "by_status": group_count("status"),
    "budget_total_usd": covered_sum("budget_usd"),
    "audience_onsite_total": covered_sum("audience_onsite"),
    "participants_total": covered_sum("participants_onsite"),
    "partners_local_total": covered_sum("partners_local"),
    "distinct_languages": distinct_langs,
    "distinct_language_count": len(distinct_langs),
    "vector_coverage": portfolio_vcov,
    "blocked_metrics": ["jobs_created_count","jobs_paid_count","partners_international_count","audience_reach_online"],
}

# ---- Notations Lab scoped case (events/team/partners) — PII stripped ----
def sheet_records(name, drop):
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = [str(c) for c in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r): continue
        rec = {hdr[i]: r[i] for i in range(len(hdr)) if hdr[i] not in drop}
        out.append(rec)
    return out, hdr

team, _ = sheet_records("team", PII_TEAM)
partners, _ = sheet_records("partners", PII_PARTNERS)
events, _ = sheet_records("events", set())
nl_disc = sorted({t.get("discipline") for t in team if t.get("discipline")})
nl_countries = sorted({p.get("country") for p in partners if p.get("country")})

notations_lab = {
    "label": "Notations Lab — flagship programme (scoped case study, NOT portfolio totals)",
    "team_size": len(team),
    "distinct_disciplines": nl_disc,
    "partner_count": len(partners),
    "partner_countries": nl_countries,
    "event_count": len(events),
    "audience_online_self_reported": 2_000_000,
    "audience_online_verified": False,
    "events": events,
    "team": team,
    "partners": partners,
}

meta = {
    "schema_version": "v7",
    "source_file": SRC.name,
    "export_date": datetime.date.today().isoformat(),
    "project_count": len(projects),
    "all_draft": True,
    "honesty_note": "Figures reflect documentation completeness, not external impact assessment. "
                    "HOSQ's A-J framework is evidence-based, not numerically scored. "
                    "Empty/zero source fields are shown as 'no data', never as measured zero.",
    "pii_stripped": True,
}

# ---- INSIGHTS (precomputed, decision-grade, honesty rules centralized) ----
def has(p, k): return p.get(k) is not None

# 1. cost-per-onsite-attendee, blended + by type (only where BOTH present & audience>0)
both = [p for p in projects if has(p,"budget_usd") and has(p,"audience_onsite") and p["audience_onsite"]>0]
eff_by_type = {}
for p in both:
    t = p["type"] or "—"; d = eff_by_type.setdefault(t, {"budget":0,"audience":0,"n":0})
    d["budget"]+=p["budget_usd"]; d["audience"]+=p["audience_onsite"]; d["n"]+=1
eff_list = sorted(({"type":t,"cost_per":round(d["budget"]/d["audience"],1),"n":d["n"],
                    "budget":d["budget"],"audience":d["audience"]} for t,d in eff_by_type.items()),
                  key=lambda x:x["cost_per"])
tb=sum(p["budget_usd"] for p in both); ta=sum(p["audience_onsite"] for p in both)
efficiency = {"n":len(both),"blended_cost_per_attendee":round(tb/ta,1) if ta else None,"by_type":eff_list}

# 2. budget concentration (Pareto)
bp = sorted([p["budget_usd"] for p in projects if has(p,"budget_usd")], reverse=True)
tot_b = sum(bp); cum=[]; run=0
for v in bp: run+=v; cum.append(round(run/tot_b,4))
concentration = {"n":len(bp),"total":tot_b,"cumulative_share":cum,
    "top":{k:round(sum(bp[:k])/tot_b,3) for k in (1,3,5,10) if k<=len(bp)}}

# 3. coverage by stream
cov_by_stream = {}
for stream in ("network","foundation"):
    ps=[p for p in projects if p["stream"]==stream]
    if ps: cov_by_stream[stream]={"n":len(ps),
        "coverage":{v:round(sum(p["vector_coverage"][v] for p in ps)/len(ps),3) for v in VEC_ORDER}}

# 4. priority vs evidence
priority_evidence=[{"vector":v,"name":vectors[v]["name"],
    "coverage":portfolio_vcov[v]["coverage"],"focus":v in {"G","A","D"}} for v in VEC_ORDER]

# 5. duration vs reach
dr=[p for p in projects if has(p,"duration_days") and has(p,"audience_onsite")]
single=[p for p in dr if p["duration_days"]<=1]; multi=[p for p in dr if p["duration_days"]>1]
duration_reach={"single":{"n":len(single),"avg":round(sum(p["audience_onsite"] for p in single)/len(single)) if single else None},
                "multi":{"n":len(multi),"avg":round(sum(p["audience_onsite"] for p in multi)/len(multi)) if multi else None}}

aggregates["insights"] = {"efficiency":efficiency,"budget_concentration":concentration,
    "coverage_by_stream":cov_by_stream,"priority_evidence":priority_evidence,"duration_reach":duration_reach}

(OUT / "manifest.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2))
(OUT / "projects.json").write_text(json.dumps(projects, ensure_ascii=False, indent=2))
(OUT / "aggregates.json").write_text(json.dumps(aggregates, ensure_ascii=False, indent=2))
(OUT / "metrics_legend.json").write_text(json.dumps({"vectors":[vectors[v] for v in VEC_ORDER]}, ensure_ascii=False, indent=2))
(OUT / "notations_lab.json").write_text(json.dumps(notations_lab, ensure_ascii=False, indent=2))

# ---- build report ----
lines = [f"HOSQ data build — {meta['export_date']}", f"source: {SRC.name}", f"projects: {len(projects)}", ""]
lines.append("KPI coverage (n_covered / n_total):")
for k in ["budget_total_usd","audience_onsite_total","participants_total","partners_local_total"]:
    a = aggregates[k]; lines.append(f"  {k}: {a['n_covered']}/{a['n_total']}  value={a['value']}")
lines.append("")
lines.append("Vector evidence coverage (portfolio avg):")
for v in VEC_ORDER:
    pv = portfolio_vcov[v]; lines.append(f"  {v} {pv['name']:<24} {pv['coverage']}  ({pv['n_fields']} fields)")
lines.append("")
lines.append(f"BLOCKED metrics (never published as totals): {aggregates['blocked_metrics']}")
lines.append(f"Notations Lab case: team={len(team)} partners={len(partners)} events={len(events)} (scoped, PII stripped)")
(OUT / "build_report.txt").write_text("\n".join(lines))
print("\n".join(lines))
print("\nWROTE:", [p.name for p in OUT.glob('*.json')], f"+ {len(projects)} project files")
